import calendar
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum
import io

from transactions.models import Income, Expense
from budgets.models import Budget
from savings.models import SavingsGoal


class MonthlyReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        now = timezone.now()
        try:
            month = int(request.query_params.get('month', now.month))
            year = int(request.query_params.get('year', now.year))
            if not (1 <= month <= 12) or year < 2000:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'error': 'Invalid month or year'}, status=status.HTTP_400_BAD_REQUEST)

        income_data = Income.objects.filter(user=user, date__month=month, date__year=year)
        expense_data = Expense.objects.filter(user=user, date__month=month, date__year=year)

        total_income = float(income_data.aggregate(t=Sum('amount'))['t'] or 0)
        total_expenses = float(expense_data.aggregate(t=Sum('amount'))['t'] or 0)

        cat_totals = {
            row['category']: float(row['t'])
            for row in expense_data.values('category').annotate(t=Sum('amount'))
        }
        category_breakdown = [
            {'category': cat, 'amount': cat_totals[cat]}
            for cat in ['food', 'transport', 'shopping', 'bills', 'health', 'education', 'entertainment', 'other']
            if cat in cat_totals
        ]

        return Response({
            'month': month,
            'year': year,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_savings': total_income - total_expenses,
            'savings_rate': round(((total_income - total_expenses) / total_income * 100) if total_income > 0 else 0, 1),
            'category_breakdown': category_breakdown,
            'income_transactions': income_data.count(),
            'expense_transactions': expense_data.count(),
        })


class ExportPDFView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            return Response({'error': 'ReportLab not installed'}, status=500)

        user = request.user
        now = timezone.now()
        try:
            month = int(request.query_params.get('month', now.month))
            year = int(request.query_params.get('year', now.year))
            if not (1 <= month <= 12) or year < 2000:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'error': 'Invalid month or year'}, status=status.HTTP_400_BAD_REQUEST)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            'Title', parent=styles['Heading1'],
            fontSize=24, textColor=colors.HexColor('#6366F1'),
            spaceAfter=20, alignment=TA_CENTER
        )

        story.append(Paragraph("AI Finance Tracker", title_style))
        story.append(Paragraph(f"Monthly Report - {calendar.month_name[month]} {year}", styles['Heading2']))
        story.append(Paragraph(f"Generated for: {user.full_name}", styles['Normal']))
        story.append(Spacer(1, 0.5 * cm))

        income_total = float(Income.objects.filter(
            user=user, date__month=month, date__year=year
        ).aggregate(t=Sum('amount'))['t'] or 0)

        expense_total = float(Expense.objects.filter(
            user=user, date__month=month, date__year=year
        ).aggregate(t=Sum('amount'))['t'] or 0)

        summary_data = [
            ['Metric', 'Amount (PKR)'],
            ['Total Income', f'₨{income_total:,.2f}'],
            ['Total Expenses', f'₨{expense_total:,.2f}'],
            ['Net Savings', f'₨{(income_total - expense_total):,.2f}'],
            ['Savings Rate', f'{round(((income_total - expense_total) / income_total * 100) if income_total > 0 else 0, 1)}%'],
        ]

        table = Table(summary_data, colWidths=[8 * cm, 8 * cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FF')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)
        story.append(Spacer(1, 1 * cm))

        story.append(Paragraph("Expense by Category", styles['Heading2']))
        expenses = Expense.objects.filter(user=user, date__month=month, date__year=year)
        cat_totals = {
            row['category']: float(row['t'])
            for row in expenses.values('category').annotate(t=Sum('amount'))
        }
        cat_data = [['Category', 'Amount (PKR)', '% of Total']]
        for cat in ['food', 'transport', 'shopping', 'bills', 'health', 'education', 'entertainment', 'other']:
            amt = cat_totals.get(cat, 0.0)
            if amt > 0:
                pct = (amt / expense_total * 100) if expense_total > 0 else 0
                cat_data.append([cat.capitalize(), f'₨{amt:,.2f}', f'{pct:.1f}%'])

        if len(cat_data) > 1:
            cat_table = Table(cat_data, colWidths=[6 * cm, 6 * cm, 4 * cm])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FF')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(cat_table)

        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="finance_report_{year}_{month:02d}.pdf"'
        return response


class ExportExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)

        user = request.user
        now = timezone.now()
        try:
            month = int(request.query_params.get('month', now.month))
            year = int(request.query_params.get('year', now.year))
            if not (1 <= month <= 12) or year < 2000:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'error': 'Invalid month or year'}, status=status.HTTP_400_BAD_REQUEST)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Finance Report"

        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')

        ws['A1'] = 'AI Finance Tracker - Monthly Report'
        ws['A1'].font = Font(bold=True, size=16, color='6366F1')
        ws['A2'] = f'Period: {calendar.month_name[month]} {year} | User: {user.full_name}'

        ws['A4'] = 'INCOME'
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'], ws['B5'], ws['C5'], ws['D5'] = 'Date', 'Source', 'Description', 'Amount (PKR)'
        for cell in ws['5:5']:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row = 6
        income_qs = Income.objects.filter(user=user, date__month=month, date__year=year)
        for inc in income_qs:
            ws.cell(row, 1, str(inc.date))
            ws.cell(row, 2, inc.source)
            ws.cell(row, 3, inc.description)
            ws.cell(row, 4, float(inc.amount))
            row += 1

        row += 2
        ws.cell(row, 1, 'EXPENSES').font = Font(bold=True, size=12)
        row += 1
        for col, val in enumerate(['Date', 'Category', 'Description', 'Amount (PKR)'], 1):
            cell = ws.cell(row, col, val)
            cell.font = header_font
            cell.fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row += 1
        expense_qs = Expense.objects.filter(user=user, date__month=month, date__year=year)
        for exp in expense_qs:
            ws.cell(row, 1, str(exp.date))
            ws.cell(row, 2, exp.category)
            ws.cell(row, 3, exp.description)
            ws.cell(row, 4, float(exp.amount))
            row += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="finance_report_{year}_{month:02d}.xlsx"'
        return response
